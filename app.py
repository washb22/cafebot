"""CafeBot - Naver Cafe Automation Flask App"""
import os
import sys
import json
import re
import random
import asyncio
import threading
import time
from queue import Queue

from flask import Flask, request, jsonify, render_template, Response
from werkzeug.utils import secure_filename
from config import FLASK_PORT, ACCOUNTS_FILE, DATA_DIR, SETTINGS_FILE, DEFAULT_IP_MODE


def _safe_image_name(original_name, num, prefix=""):
    """업로드 이미지 저장명 생성 — 확장자 보존 보장.
    한글/특수문자가 포함된 원본 이름은 secure_filename 이 확장자까지 날리기 때문에
    확장자를 별도로 뽑아 보존한다.
    """
    original_name = original_name or ""
    ext = os.path.splitext(original_name)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"):
        ext = ".jpg"  # 기본값 — 네이버가 확장자 없으면 거부
    # 원본 이름의 sanitize 결과가 비거나 확장자 잃었으면 img{num}{ext} 로 대체
    base = os.path.splitext(original_name)[0]
    clean = secure_filename(base)
    if not clean:
        clean = f"img{num}"
    ts = str(int(time.time() * 1000))
    return f"{ts}_{prefix}{num}_{clean}{ext}"

# Ensure data + uploads dir
os.makedirs(DATA_DIR, exist_ok=True)
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# frozen 빌드 시 템플릿 경로
if getattr(sys, 'frozen', False):
    _template_dir = os.path.join(sys._MEIPASS, 'templates')
else:
    _template_dir = os.path.join(os.path.dirname(__file__), 'templates')

app = Flask(__name__, template_folder=_template_dir)

# ── State ──
from collections import deque
log_queue = Queue()
log_history = deque(maxlen=50000)  # 로그 전체 보존 (다운로드용)
task_running = False
stop_event = threading.Event()


def _emit_log(msg):
    """log_fn 통합 헬퍼: queue(SSE) + history(다운로드) 동시 적재."""
    ts = time.strftime("%H:%M:%S")
    entry = {"type": "log", "time": ts, "message": msg}
    log_queue.put(entry)
    log_history.append(entry)


# ── Account management ──

def load_accounts():
    """accounts.json 로드 + 구버전 schema 자동 마이그레이션.
    구: {"main": {...} or null, "commenters": [...]}
    신: {"mains": [{id, pw, label="글 1"}, ...], "commenters": [...]}
    구 main 필드는 읽을 때 mains[0] 로 흡수되고, 저장할 때 함께 기록(호환).
    """
    if os.path.exists(ACCOUNTS_FILE):
        with open(ACCOUNTS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = {"mains": [], "commenters": []}

    # 마이그레이션: main 단수 → mains 배열
    if "mains" not in data:
        data["mains"] = []
    if data.get("main"):
        m = data["main"]
        # 이미 mains 에 같은 id 있으면 스킵
        if not any(x.get("id") == m.get("id") for x in data["mains"]):
            data["mains"].insert(0, {
                "id": m["id"],
                "pw": m["pw"],
                "label": m.get("label") or f"글 {len(data['mains']) + 1}",
            })
    # main 필드 제거 (신 schema)
    data.pop("main", None)
    if "commenters" not in data:
        data["commenters"] = []
    return data


def load_settings():
    """설정 파일(ip_mode 등) 로드. 파일이 없거나 깨져있으면 기본값."""
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = {}
    else:
        data = {}
    mode = (data.get("ip_mode") or DEFAULT_IP_MODE).lower()
    if mode not in ("proxy", "adb"):
        mode = DEFAULT_IP_MODE
    data["ip_mode"] = mode
    return data


def save_settings(data):
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_ip_mode():
    return load_settings().get("ip_mode", DEFAULT_IP_MODE)


def save_accounts(data):
    # main 필드가 올라오면 mains[0] 로 변환
    if "mains" not in data:
        data["mains"] = []
    if data.get("main"):
        m = data["main"]
        if not any(x.get("id") == m.get("id") for x in data["mains"]):
            data["mains"].insert(0, {
                "id": m["id"],
                "pw": m["pw"],
                "label": m.get("label") or "글 1",
            })
        data.pop("main", None)
    with open(ACCOUNTS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _resolve_main(accounts, main_id=None):
    """mains 배열에서 main_id 로 계정 선택. 미지정 시 첫 번째."""
    mains = accounts.get("mains", [])
    if not mains:
        return None
    if main_id:
        for m in mains:
            if m.get("id") == main_id:
                return m
        return None  # main_id 지정됐는데 없으면 실패
    return mains[0]


# ── Routes ──

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/settings", methods=["GET"])
def get_settings_api():
    return jsonify(load_settings())


@app.route("/api/settings", methods=["POST"])
def update_settings_api():
    data = request.json or {}
    current = load_settings()
    # 허용 필드만 갱신
    if "ip_mode" in data:
        mode = str(data.get("ip_mode") or "").lower()
        if mode not in ("proxy", "adb"):
            return jsonify({"error": "ip_mode 는 proxy 또는 adb 여야 합니다"}), 400
        current["ip_mode"] = mode
    save_settings(current)
    return jsonify({"success": True, "settings": current})


@app.route("/api/accounts", methods=["GET"])
def get_accounts():
    return jsonify(load_accounts())


@app.route("/api/accounts", methods=["POST"])
def update_accounts():
    data = request.json
    save_accounts(data)
    return jsonify({"success": True})


@app.route("/api/accounts/proxy", methods=["POST"])
def update_account_proxy():
    """계정 1개의 프록시 매핑만 수정. body: {type:'main'|'commenter', index, proxy}"""
    data = request.json or {}
    t = data.get("type")
    idx = data.get("index")
    proxy = (data.get("proxy") or "").strip()
    if t not in ("main", "commenter") or not isinstance(idx, int):
        return jsonify({"error": "잘못된 요청"}), 400
    accounts = load_accounts()
    key = "mains" if t == "main" else "commenters"
    arr = accounts.get(key, [])
    if idx < 0 or idx >= len(arr):
        return jsonify({"error": "인덱스 범위 초과"}), 400
    arr[idx]["proxy"] = proxy
    save_accounts(accounts)
    return jsonify({"success": True, "proxy": proxy})


@app.route("/api/accounts/proxy/bulk", methods=["POST"])
def update_proxies_bulk():
    """여러 계정 프록시 일괄 매핑. body: {commenters: ["host:port", ...], mains: [...]}
    배열 길이만큼 순서대로 매핑. 빈 문자열은 스킵."""
    data = request.json or {}
    accounts = load_accounts()
    for key in ("commenters", "mains"):
        vals = data.get(key)
        if not isinstance(vals, list):
            continue
        arr = accounts.get(key, [])
        for i, v in enumerate(vals):
            if i >= len(arr):
                break
            if v is None:
                continue
            arr[i]["proxy"] = str(v).strip()
    save_accounts(accounts)
    return jsonify({"success": True})


@app.route("/api/accounts/excel", methods=["POST"])
def upload_accounts_excel():
    """엑셀로 계정 일괄 등록/갱신.
    컬럼: A=ID, B=PW, C=역할(글작성자/댓글), D=프록시(host:port), E=라벨(선택)
    기존 ID 와 같으면 PW/프록시/라벨 업데이트 (삭제는 안 함).
    """
    import openpyxl
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "파일이 없습니다"}), 400

    try:
        wb = openpyxl.load_workbook(f, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return jsonify({"error": f"엑셀 읽기 오류: {e}"}), 400

    if not rows:
        return jsonify({"error": "빈 파일입니다"}), 400

    # 헤더 판별
    first = [str(c or "").strip().lower() for c in rows[0]]
    header_keywords = ["id", "아이디", "pw", "비밀번호", "password", "역할", "role", "프록시", "proxy", "라벨", "label"]
    if any(k in first for k in header_keywords):
        rows = rows[1:]

    accounts = load_accounts()
    added_main = 0
    added_comment = 0
    updated_main = 0
    updated_comment = 0

    for row in rows:
        if not row or len(row) < 2:
            continue
        acc_id = str(row[0] or "").strip()
        acc_pw = str(row[1] or "").strip()
        if not acc_id or not acc_pw:
            continue

        role = "commenter"
        if len(row) >= 3 and row[2]:
            role_text = str(row[2]).strip()
            low = role_text.lower()
            # 댓글 계정은 '댓글' 문자열이 포함됨 → 우선 커멘터로 고정
            if "댓글" in role_text or "reply" in low or "comment" in low:
                role = "commenter"
            elif "글작성" in role_text or "작성자" in role_text or "main" in low:
                role = "main"

        proxy_val = ""
        if len(row) >= 4 and row[3]:
            proxy_val = str(row[3]).strip()

        custom_label = ""
        if len(row) >= 5 and row[4]:
            custom_label = str(row[4]).strip()

        target_key = "mains" if role == "main" else "commenters"
        arr = accounts[target_key]

        existing = next((x for x in arr if x.get("id") == acc_id), None)
        if existing:
            existing["pw"] = acc_pw
            if proxy_val or "proxy" in existing:
                existing["proxy"] = proxy_val
            if custom_label:
                existing["label"] = custom_label
            if role == "main":
                updated_main += 1
            else:
                updated_comment += 1
        else:
            default_label = (
                f"글 {len(accounts['mains']) + 1}" if role == "main"
                else f"댓글 {len(accounts['commenters']) + 1}"
            )
            arr.append({
                "id": acc_id,
                "pw": acc_pw,
                "label": custom_label or default_label,
                "proxy": proxy_val,
            })
            if role == "main":
                added_main += 1
            else:
                added_comment += 1

    save_accounts(accounts)
    parts = []
    if added_main or added_comment:
        parts.append(f"추가: 글작성자 {added_main}개 / 댓글 {added_comment}개")
    if updated_main or updated_comment:
        parts.append(f"갱신: 글작성자 {updated_main}개 / 댓글 {updated_comment}개")
    msg = ", ".join(parts) or "변경사항 없음"
    return jsonify({
        "success": True,
        "message": msg,
        "added_main": added_main,
        "added_comment": added_comment,
        "updated_main": updated_main,
        "updated_comment": updated_comment,
    })


@app.route("/api/accounts/excel/export", methods=["GET"])
def export_accounts_excel():
    """현재 계정 + 프록시 매핑 전체를 엑셀로 다운로드.
    컬럼: ID / PW / 역할 / 프록시 / 라벨
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO

    accounts = load_accounts()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "cafebot accounts"

    headers = ["ID", "PW", "역할", "프록시", "라벨"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor="E0E4FF")

    for m in accounts.get("mains", []):
        ws.append([m.get("id",""), m.get("pw",""), "글작성자", m.get("proxy",""), m.get("label","")])
    for c in accounts.get("commenters", []):
        ws.append([c.get("id",""), c.get("pw",""), "댓글", c.get("proxy",""), c.get("label","")])

    # 컬럼 폭 자동
    widths = [16, 14, 10, 28, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = time.strftime("%Y%m%d_%H%M%S")
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="cafebot_accounts_{ts}.xlsx"'
        },
    )


@app.route("/api/comment-only/run", methods=["POST"])
def run_comment_only():
    """기존 게시글에 댓글만 작업 (글 작성/수정 없음)"""
    global task_running
    if task_running:
        return jsonify({"error": "이미 작업이 실행 중입니다"}), 400

    from modules.txt_parser import parse_scenario_text

    # multipart 또는 json 모두 수용
    if request.content_type and request.content_type.startswith("multipart/form-data"):
        data = {
            "text": request.form.get("text", ""),
            "post_url": request.form.get("post_url", "").strip(),
            "main_id": request.form.get("main_id", "").strip(),
        }
    else:
        data = request.json or {}

    post_url = data.get("post_url", "").strip()
    if not post_url:
        return jsonify({"error": "게시글 URL을 입력해주세요"}), 400

    raw_text = data.get("text", "")
    if not raw_text:
        return jsonify({"error": "시나리오 텍스트가 없습니다"}), 400

    try:
        parsed = parse_scenario_text(raw_text)
    except Exception as e:
        return jsonify({"error": f"파싱 오류: {e}"}), 400

    accounts = load_accounts()
    main_acc = _resolve_main(accounts, data.get("main_id"))
    if not main_acc:
        return jsonify({"error": "메인(글 작성자) 계정이 선택되지 않았거나 존재하지 않습니다"}), 400

    commenter_map = {}
    for c in accounts.get("commenters", []):
        label = c.get("label", "")
        m = re.search(r'\d+', label)
        if m:
            commenter_map[int(m.group())] = c

    ip_mode = get_ip_mode()
    exec_actions, spare_commenters, err = build_shuffled_exec_actions(parsed, commenter_map, main_acc, ip_mode=ip_mode)
    if err:
        return jsonify({"error": err}), 400

    task = {
        "mode": "comment_only",
        "post_url": post_url,
        "cafe_url": "",
        "board_name": "",
        "title": "",
        "body": "",
        "main_account": main_acc,
        "comments": [],
        "replies": [],
        "scenario": exec_actions,
        "spare_commenters": spare_commenters,
        "image_map": {},
        "ip_mode": ip_mode,
    }

    stop_event.clear()
    task_running = True

    def log_fn(msg):
        _emit_log(msg)

    def run_in_thread():
        global task_running
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            from modules.task_runner import run_task
            result = loop.run_until_complete(run_task(task, log_fn, stop_event))
            log_queue.put({"type": "done", "result": result})
        except Exception as e:
            log_queue.put({"type": "error", "message": str(e)})
        finally:
            task_running = False

    threading.Thread(target=run_in_thread, daemon=True).start()
    return jsonify({"success": True, "message": f"댓글 작업 시작 ({len(exec_actions)}개 액션)"})


@app.route("/api/proxy/healthcheck", methods=["POST"])
def proxy_healthcheck():
    """전체 계정 프록시 헬스체크 (네이버 접근 없음, 순수 IP 확인 서비스만)."""
    global task_running
    if task_running:
        return jsonify({"error": "작업 실행 중에는 헬스체크할 수 없습니다"}), 400

    accounts = load_accounts()

    def log_fn(msg):
        _emit_log(f"[헬스체크] {msg}")

    try:
        from modules.proxy_health import check_all_proxies
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            results = loop.run_until_complete(check_all_proxies(accounts, concurrency=5, log_fn=log_fn))
        finally:
            loop.close()
    except Exception as e:
        return jsonify({"error": f"헬스체크 오류: {e}"}), 500

    # 요약 계산
    all_rows = (results.get("mains") or []) + (results.get("commenters") or [])
    summary = {
        "total": len(all_rows),
        "ok": sum(1 for r in all_rows if r["status"] == "ok"),
        "mismatch": sum(1 for r in all_rows if r["status"] == "mismatch"),
        "unreachable": sum(1 for r in all_rows if r["status"] == "unreachable"),
        "no_proxy": sum(1 for r in all_rows if r["status"] == "no_proxy"),
        "error": sum(1 for r in all_rows if r["status"] == "error"),
    }
    return jsonify({"success": True, "summary": summary, "results": results})


@app.route("/api/adb/status")
def adb_status():
    from modules.adb_network import is_device_connected, get_current_ip
    return jsonify({
        "connected": is_device_connected(),
        "ip": get_current_ip()
    })


@app.route("/api/logs/download")
def logs_download():
    """로그 히스토리 전체를 .txt 파일로 다운로드 (브라우저 환경)."""
    lines = [f"{e.get('time','')}\t{e.get('message','')}" for e in log_history]
    header = [
        f"# CafeBot 실행 로그 (서버 히스토리)",
        f"# 생성: {time.strftime('%Y-%m-%d %H:%M:%S')}",
        f"# 총 {len(lines)} 줄",
        "",
    ]
    body = "\n".join(header + lines)
    ts = time.strftime("%Y%m%d_%H%M%S")
    return Response(
        body.encode("utf-8"),
        mimetype="text/plain; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="cafebot_log_{ts}.txt"',
        },
    )


@app.route("/api/logs/save_to_downloads", methods=["POST"])
def logs_save_to_downloads():
    """WebView2 다운로드 막힘 우회: 서버가 직접 ~/Downloads 에 파일 저장."""
    lines = [f"{e.get('time','')}\t{e.get('message','')}" for e in log_history]
    if not lines:
        return jsonify({"success": False, "error": "저장할 로그가 없습니다"}), 400
    header = [
        f"# CafeBot 실행 로그",
        f"# 생성: {time.strftime('%Y-%m-%d %H:%M:%S')}",
        f"# 총 {len(lines)} 줄",
        "",
    ]
    text = "\n".join(header + lines)
    downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
    try:
        os.makedirs(downloads_dir, exist_ok=True)
    except Exception:
        downloads_dir = os.path.expanduser("~")
    ts = time.strftime("%Y%m%d_%H%M%S")
    fname = f"cafebot_log_{ts}.txt"
    fpath = os.path.join(downloads_dir, fname)
    try:
        with open(fpath, "w", encoding="utf-8") as f:
            f.write(text)
        return jsonify({"success": True, "path": fpath, "count": len(lines)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/logs/text")
def logs_text():
    """로그 히스토리를 JSON 으로 반환 (클립보드 복사용)."""
    lines = [f"{e.get('time','')}\t{e.get('message','')}" for e in log_history]
    return jsonify({"text": "\n".join(lines), "count": len(lines)})


@app.route("/api/logs/clear", methods=["POST"])
def logs_clear():
    log_history.clear()
    return jsonify({"success": True})


@app.route("/api/logs/stream")
def log_stream():
    """SSE endpoint for real-time logs."""
    def generate():
        while True:
            try:
                msg = log_queue.get(timeout=30)
                yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
            except Exception:
                yield f"data: {json.dumps({'type': 'ping'})}\n\n"
    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/api/tasks/run", methods=["POST"])
def run_task_api():
    global task_running
    if task_running:
        return jsonify({"error": "이미 작업이 실행 중입니다"}), 400

    task_data = request.json
    accounts = load_accounts()

    # Build task object
    main_acc = _resolve_main(accounts, task_data.get("main_id"))
    if not main_acc:
        return jsonify({"error": "메인(글 작성자) 계정이 선택되지 않았거나 존재하지 않습니다"}), 400

    commenters = accounts.get("commenters", [])
    commenter_map = {c["id"]: c for c in commenters}

    comments = []
    for c in task_data.get("comments", []):
        acc = commenter_map.get(c["account_id"])
        if acc:
            comments.append({"account": acc, "text": c["text"]})

    replies = []
    for r in task_data.get("replies", []):
        replies.append({"to_index": r["to_index"], "text": r["text"]})

    task = {
        "mode": task_data.get("mode", "new"),
        "cafe_url": task_data.get("cafe_url", ""),
        "post_url": task_data.get("post_url", ""),
        "board_name": task_data.get("board_name", ""),
        "title": task_data.get("title", ""),
        "body": task_data.get("body", ""),
        "main_account": main_acc,
        "comments": comments,
        "replies": replies,
        "ip_mode": get_ip_mode(),
    }

    # Run in background thread
    stop_event.clear()
    task_running = True

    def log_fn(msg):
        _emit_log(msg)

    def run_in_thread():
        global task_running
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            from modules.task_runner import run_task
            result = loop.run_until_complete(run_task(task, log_fn, stop_event))
            log_queue.put({"type": "done", "result": result})
        except Exception as e:
            log_queue.put({"type": "error", "message": str(e)})
        finally:
            task_running = False

    t = threading.Thread(target=run_in_thread, daemon=True)
    t.start()

    return jsonify({"success": True, "message": "작업 시작됨"})


def build_shuffled_exec_actions(parsed, commenter_map, main_acc, shuffle_label="", ip_mode=None):
    """parsed 시나리오 → 실행용 actions 생성 + 계정 셔플.

    텍스트 파일의 댓글 순서는 그대로 유지하되, 역할번호(댓글1/댓글2/...)를
    실제 commenter 계정에 랜덤 배정한다. 같은 역할번호는 같은 계정으로 매핑 유지
    → 대댓글 대화 일관성 보존. to_index도 원본 그대로 유지.

    ip_mode: "proxy" 면 프록시 설정된 계정만 셔플 풀에 포함.
             "adb" 면 모든 등록 계정을 풀에 포함(IP 는 비행기모드 토글로 바꿈).
             None 이면 settings 에서 읽음.

    Returns: (exec_actions, spare_commenters, error_message_or_None)
      spare_commenters: 시나리오에 배정되지 않은 여분 댓글러 풀.
        proxy 모드에서 task_runner 가 unreachable 발생 시 교체용으로 사용.
    """
    if ip_mode is None:
        ip_mode = get_ip_mode()

    needed_nums = list(parsed["commenter_nums"])
    total = len(commenter_map)

    if ip_mode == "proxy":
        # 프록시 설정된 commenter 만 셔플 풀에 포함 (IP 미변경 방지)
        available = [k for k, v in commenter_map.items() if (v.get("proxy") or "").strip()]
        if len(available) < len(needed_nums):
            return None, [], (
                f"프록시 설정된 commenter 부족: 필요 {len(needed_nums)}개 / "
                f"프록시 설정됨 {len(available)}개 (전체 {total}개). "
                "계정 관리 → 프록시 매핑에서 host:port 입력 후 저장하거나, "
                "우측 상단에서 ADB 모드로 전환하세요."
            )
        if not (main_acc.get("proxy") or "").strip():
            return None, [], (
                f"글 작성자({main_acc.get('id','?')}) 에 프록시가 설정되지 않았습니다. "
                "계정 관리 → 프록시 매핑에서 설정하거나, ADB 모드로 전환하세요."
            )
    else:
        # ADB 모드: 모든 등록 계정이 풀 (프록시 필드 무시)
        available = list(commenter_map.keys())
        if len(available) < len(needed_nums):
            return None, [], (
                f"commenter 계정 부족: 필요 {len(needed_nums)}개 / 등록 {total}개."
            )

    # 역할번호 → 실제 계정 랜덤 매핑 (프록시 있는 계정만)
    picked = random.sample(available, len(needed_nums))
    role_to_account = {
        role: commenter_map[picked[i]]
        for i, role in enumerate(needed_nums)
    }

    # 원본 순서 그대로 exec_actions 생성 (계정만 셔플된 매핑 사용)
    exec_actions = []
    for act in parsed["actions"]:
        if act["action"] == "comment":
            acc = role_to_account.get(act["commenter_num"])
            if not acc:
                return None, [], f"댓글 {act['commenter_num']} 계정이 등록되지 않음"
            exec_actions.append({"action": "comment", "account": acc, "text": act["text"]})
        elif act["action"] == "reply":
            if act.get("is_main"):
                acc = main_acc
            else:
                acc = role_to_account.get(act["commenter_num"])
                if not acc:
                    return None, [], f"ㄴ 댓글 {act['commenter_num']} 계정이 등록되지 않음"
            exec_actions.append({
                "action": "reply", "account": acc,
                "to_index": act["to_index"], "text": act["text"],
                "is_main": bool(act.get("is_main")),  # 2-브라우저 구조 플래그 (task_runner 가 사용)
            })

    # 여분 댓글러 풀 (배정되지 않은 계정 — proxy 모드에서 unreachable 발생 시 교체용)
    picked_set = set(picked)
    spare_commenters = [commenter_map[k] for k in available if k not in picked_set]
    random.shuffle(spare_commenters)

    # 디버그용 요약
    mapping_summary = ", ".join(
        f"역할{role}→{role_to_account[role].get('id','?')}"
        for role in needed_nums
    )
    _emit_log(f"[계정셔플{shuffle_label}] {mapping_summary}")
    if spare_commenters:
        _emit_log(f"[교체풀{shuffle_label}] 여분 댓글러 {len(spare_commenters)}개 (프록시 unreachable 시 자동 교체)")

    return exec_actions, spare_commenters, None


@app.route("/api/scenario/parse", methods=["POST"])
def parse_scenario():
    """업로드된 txt 파일을 파싱해 프리뷰 반환"""
    from modules.txt_parser import parse_scenario_text
    if 'file' in request.files:
        raw = request.files['file'].read().decode('utf-8', errors='replace')
    else:
        raw = (request.json or {}).get('text', '')
    try:
        parsed = parse_scenario_text(raw)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    accounts = load_accounts()
    commenter_map = {}
    for c in accounts.get("commenters", []):
        label = c.get("label", "")
        m = re.search(r'\d+', label)
        if m:
            commenter_map[int(m.group())] = c

    missing = [n for n in parsed["commenter_nums"] if n not in commenter_map]
    mapped = {n: commenter_map[n]["id"] for n in parsed["commenter_nums"] if n in commenter_map}

    # 프록시 가용성 요약 (셔플 풀 = 프록시 설정된 commenter 만)
    proxy_ready_count = sum(
        1 for c in accounts.get("commenters", []) if (c.get("proxy") or "").strip()
    )
    needed_count = len(parsed["commenter_nums"])
    proxy_ok = proxy_ready_count >= needed_count

    return jsonify({
        "title": parsed["title"],
        "body": parsed["body"],
        "actions": parsed["actions"],
        "commenter_nums": parsed["commenter_nums"],
        "image_nums": parsed.get("image_nums", []),
        "mapped_accounts": mapped,
        "missing_nums": missing,
        "proxy_ready_count": proxy_ready_count,
        "proxy_needed_count": needed_count,
        "proxy_ok": proxy_ok,
    })


@app.route("/api/scenario/run", methods=["POST"])
def run_scenario():
    """시나리오 기반 실행 (txt에서 파싱한 actions + 카페 URL).
    multipart/form-data 지원: text, cafe_url, main_id, board_name + image_1, image_2, ...
    JSON 도 호환 (이미지 없을 때).
    """
    global task_running
    if task_running:
        return jsonify({"error": "이미 작업이 실행 중입니다"}), 400

    from modules.txt_parser import parse_scenario_text

    # multipart 또는 json 모두 수용
    image_map = {}
    if request.content_type and request.content_type.startswith("multipart/form-data"):
        data = {
            "text": request.form.get("text", ""),
            "cafe_url": request.form.get("cafe_url", ""),
            "board_name": request.form.get("board_name", ""),
            "main_id": request.form.get("main_id", ""),
        }
        # image_1, image_2, ... 저장
        for key, f in request.files.items():
            if not key.startswith("image_") or not f or not f.filename:
                continue
            try:
                num = int(key.split("_", 1)[1])
            except ValueError:
                continue
            fname = _safe_image_name(f.filename, num)
            path = os.path.join(UPLOAD_DIR, fname)
            f.save(path)
            image_map[num] = path
    else:
        data = request.json or {}

    raw_text = data.get("text", "")
    try:
        parsed = parse_scenario_text(raw_text)
    except Exception as e:
        return jsonify({"error": f"파싱 오류: {e}"}), 400

    accounts = load_accounts()
    main_acc = _resolve_main(accounts, data.get("main_id"))
    if not main_acc:
        return jsonify({"error": "메인(글 작성자) 계정이 선택되지 않았거나 존재하지 않습니다"}), 400

    commenter_map = {}
    for c in accounts.get("commenters", []):
        label = c.get("label", "")
        m = re.search(r'\d+', label)
        if m:
            commenter_map[int(m.group())] = c

    ip_mode = get_ip_mode()
    # 시나리오 actions → 실행용 actions (계정 셔플 + 그룹 순서 셔플)
    exec_actions, spare_commenters, err = build_shuffled_exec_actions(parsed, commenter_map, main_acc, ip_mode=ip_mode)
    if err:
        return jsonify({"error": err}), 400

    task = {
        "mode": "new",
        "cafe_url": data.get("cafe_url", ""),
        "board_name": data.get("board_name", ""),
        "title": parsed["title"],
        "body": parsed["body"],
        "main_account": main_acc,
        "comments": [],
        "replies": [],
        "scenario": exec_actions,
        "spare_commenters": spare_commenters,
        "image_map": image_map,
        "ip_mode": ip_mode,
    }

    stop_event.clear()
    task_running = True

    def log_fn(msg):
        _emit_log(msg)

    def run_in_thread():
        global task_running
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            from modules.task_runner import run_task
            result = loop.run_until_complete(run_task(task, log_fn, stop_event))
            log_queue.put({"type": "done", "result": result})
        except Exception as e:
            log_queue.put({"type": "error", "message": str(e)})
        finally:
            task_running = False

    threading.Thread(target=run_in_thread, daemon=True).start()
    return jsonify({"success": True, "message": f"시나리오 실행 시작 ({len(exec_actions)}개 액션)"})


@app.route("/api/queue/run", methods=["POST"])
def run_queue():
    """여러 시나리오를 한 번에 받아 순차 실행.

    multipart/form-data:
        count=<N>
        s0_text, s0_cafe_url, s0_main_id, s0_board_name, s0_image_1, s0_image_2, ...
        s1_text, ...
    """
    global task_running
    if task_running:
        return jsonify({"error": "이미 작업이 실행 중입니다"}), 400

    from modules.txt_parser import parse_scenario_text

    try:
        count = int(request.form.get("count", "0"))
    except ValueError:
        count = 0
    if count <= 0:
        return jsonify({"error": "큐에 작업이 없습니다"}), 400

    accounts = load_accounts()
    commenter_map = {}
    for c in accounts.get("commenters", []):
        label = c.get("label", "")
        m = re.search(r'\d+', label)
        if m:
            commenter_map[int(m.group())] = c

    batch_ip_mode = get_ip_mode()
    tasks = []
    for i in range(count):
        prefix = f"s{i}_"
        raw_text = request.form.get(prefix + "text", "")
        mode = request.form.get(prefix + "mode", "new").strip() or "new"
        cafe_url = request.form.get(prefix + "cafe_url", "").strip()
        post_url_field = request.form.get(prefix + "post_url", "").strip()
        main_id = request.form.get(prefix + "main_id", "").strip()
        board_name = request.form.get(prefix + "board_name", "").strip()

        if not raw_text or not main_id:
            return jsonify({"error": f"작업 #{i+1}: text/main_id 누락"}), 400
        if mode == "new" and not cafe_url:
            return jsonify({"error": f"작업 #{i+1}: 새 글 작성은 카페 URL 필요"}), 400
        if mode == "edit" and not post_url_field:
            return jsonify({"error": f"작업 #{i+1}: 수정 모드는 기존 글 URL 필요"}), 400

        try:
            parsed = parse_scenario_text(raw_text)
        except Exception as e:
            return jsonify({"error": f"작업 #{i+1} 파싱 오류: {e}"}), 400

        main_acc = _resolve_main(accounts, main_id)
        if not main_acc:
            return jsonify({"error": f"작업 #{i+1}: 글 작성자 '{main_id}' 없음"}), 400

        # 이미지 파일 저장
        image_map = {}
        for key, f in request.files.items():
            if not key.startswith(prefix + "image_"):
                continue
            try:
                num = int(key[len(prefix) + len("image_"):])
            except ValueError:
                continue
            if not f or not f.filename:
                continue
            fname = _safe_image_name(f.filename, num, prefix=f"q{i}_")
            path = os.path.join(UPLOAD_DIR, fname)
            f.save(path)
            image_map[num] = path

        # 시나리오 actions → 실행용 (계정 셔플 + 그룹 순서 셔플, 글마다 독립 재셔플)
        exec_actions, spare_commenters, err = build_shuffled_exec_actions(
            parsed, commenter_map, main_acc, shuffle_label=f" 글{i+1}", ip_mode=batch_ip_mode
        )
        if err:
            return jsonify({"error": f"작업 #{i+1}: {err}"}), 400

        tasks.append({
            "mode": mode,
            "cafe_url": cafe_url,
            "post_url": post_url_field,
            "board_name": board_name,
            "title": parsed["title"],
            "body": parsed["body"],
            "main_account": main_acc,
            "comments": [],
            "replies": [],
            "scenario": exec_actions,
            "spare_commenters": spare_commenters,
            "image_map": image_map,
            "ip_mode": batch_ip_mode,
        })

    stop_event.clear()
    task_running = True

    def log_fn(msg):
        _emit_log(msg)

    def run_in_thread():
        global task_running
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            from modules.task_runner import run_batch
            result = loop.run_until_complete(run_batch(tasks, log_fn, stop_event))
            log_queue.put({"type": "done", "result": result})
        except Exception as e:
            log_queue.put({"type": "error", "message": str(e)})
        finally:
            task_running = False

    threading.Thread(target=run_in_thread, daemon=True).start()
    return jsonify({"success": True, "message": f"배치 시작: {count}개 작업"})


@app.route("/api/tasks/stop", methods=["POST"])
def stop_task():
    global task_running
    if task_running:
        stop_event.set()
        return jsonify({"success": True, "message": "중단 요청됨"})
    return jsonify({"success": False, "message": "실행 중인 작업 없음"})


@app.route("/api/tasks/status")
def task_status():
    return jsonify({"running": task_running})


# ── License API ──

@app.route("/api/license/status")
def license_status():
    from modules.license import get_status
    return jsonify(get_status())

@app.route("/api/license/activate", methods=["POST"])
def license_activate():
    data = request.json or {}
    key = data.get("license_key", "").strip()
    if not key:
        return jsonify({"success": False, "error": "라이선스 키를 입력해주세요"}), 400
    from modules.license import activate
    ok, msg, days = activate(key)
    return jsonify({"success": ok, "message": msg, "days_remaining": days})


def run_server():
    """PyWebView 등 외부 진입점에서 호출하는 서버 시작 함수."""
    app.run(host="127.0.0.1", port=FLASK_PORT, threaded=True, use_reloader=False)


if __name__ == "__main__":
    # debug=False: 파일 수정 시 자동 재시작으로 실행 중인 작업이 죽는 걸 방지
    app.run(debug=False, port=FLASK_PORT, threaded=True, use_reloader=False)
